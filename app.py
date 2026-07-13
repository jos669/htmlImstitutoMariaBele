from datetime import datetime, timedelta
from flask import Flask, render_template, request, jsonify, session, send_file
from functools import wraps
import csv
import io
import os
import secrets
import time
import re
import socket
import subprocess
import threading
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_DISPONIBLE = True
except ImportError:
    OPENPYXL_DISPONIBLE = False
try:
    import qrcode
    from io import BytesIO
    import base64
    QR_DISPONIBLE = True
except ImportError:
    QR_DISPONIBLE = False
from database import (
    get_connection,
    validar_usuario,
    obtener_usuario_por_cui,
    registrar_consentimiento,
    obtener_todos_usuarios,
    obtener_tipos,
    crear_tipo,
    actualizar_tipo,
    eliminar_tipo,
    obtener_papeleria_estudiante,
    actualizar_papeleria,
    actualizar_papeleria_batch,
    buscar_papeleria_por_estado,
    obtener_stats_por_grado,
    obtener_estudiantes_incompletos,
    buscar_estudiantes,
    crear_usuario,
    actualizar_usuario,
    eliminar_usuario,
    registrar_log,
    obtener_grados_secciones,
    init_db,
    ANIOS,
    ESTADOS,
    CATEGORIAS,
    obtener_token_verificacion,
    verificar_por_token
)
from importador import parsear_archivo, normalizar_registros
from expediente import (
    init_tablas_expediente,
    crear_expediente,
    obtener_expediente,
    actualizar_expediente,
    cambiar_estado_expediente,
    registrar_verificacion,
    obtener_movimientos,
    obtener_extraviados,
    obtener_expedientes_por_grado,
    ESTADOS_FISICOS
)

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', secrets.token_hex(32))
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=15)

# ── CSRF Protection ──
def generar_csrf_token():
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(32)
    return session['csrf_token']

def csrf_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if request.method in ('POST', 'PUT', 'DELETE', 'PATCH'):
            token = request.headers.get('X-CSRF-Token') or request.form.get('_csrf_token')
            if not token or token != session.get('csrf_token'):
                return jsonify({'error': 'CSRF token inválido o faltante. Recargue la página.'}), 403
        return f(*args, **kwargs)
    return decorated

app.jinja_env.globals['csrf_token'] = generar_csrf_token

# ── Rate Limiting (login) ──
_attempts = {}
def check_rate_limit(ip):
    now = time.time()
    if ip in _attempts:
        count, window_start = _attempts[ip]
        if now - window_start > 900:  # 15 min window
            _attempts[ip] = [1, now]
            return True
        if count >= 5:
            return False
        _attempts[ip] = [count + 1, window_start]
    else:
        _attempts[ip] = [1, now]
    return True

def reset_rate_limit(ip):
    _attempts.pop(ip, None)

init_db()
init_tablas_expediente()

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'cui' not in session:
            return jsonify({'error': 'No autenticado. Por favor inicie sesión.'}), 401
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'cui' not in session or session.get('rol') != 'admin':
            return jsonify({'error': 'Acceso no autorizado.'}), 403
        return f(*args, **kwargs)
    return decorated_function

@app.route('/')
def index_view():
    return render_template('index.html')

@app.route('/admin')
def admin_view():
    if 'cui' not in session or session.get('rol') != 'admin':
        return render_template('admin_login.html')
    return render_template('admin.html')

@app.route('/api/login', methods=['POST'])
@csrf_required
def api_login():
    ip = request.remote_addr or 'unknown'
    if not check_rate_limit(ip):
        return jsonify({'error': 'Demasiados intentos. Intente de nuevo en 15 minutos.'}), 429

    data = request.get_json() or {}
    cui = data.get('cui')
    password = data.get('password')
    acepto_checkbox = data.get('consentimiento')

    if not cui:
        return jsonify({'error': 'CUI es obligatorio.'}), 400

    user = obtener_usuario_por_cui(cui)
    if not user:
        registrar_log(cui if cui else "desconocido", "login_fallido", "Intento de inicio de sesion fallido", request.remote_addr)
        return jsonify({'error': 'CUI no registrado en el sistema.'}), 401

    if user['rol'] == 'admin':
        if not password:
            return jsonify({'error': 'Contraseña requerida para acceso administrativo.'}), 400
        user = validar_usuario(cui, password)
        if not user:
            registrar_log(cui, "login_fallido", "Intento de inicio de sesion admin fallido", request.remote_addr)
            return jsonify({'error': 'Credenciales administrativas incorrectas.'}), 401

    if user['rol'] == 'estudiante':
        if not user['acepto_consentimiento']:
            if acepto_checkbox:
                registrar_consentimiento(cui)
                user['acepto_consentimiento'] = 1
                registrar_log(cui, 'consentimiento_aceptado', 'Acepto expresamente los terminos de consentimiento de datos', request.remote_addr)
            else:
                return jsonify({
                    'consentimiento_requerido': True,
                    'mensaje': 'Debe leer y aceptar el consentimiento de datos escolares para ingresar al portal.'
                }), 200

    reset_rate_limit(ip)
    session.permanent = True
    session['cui'] = user['cui']
    session['nombre'] = user['nombre_completo']
    session['rol'] = user['rol']
    session['grado'] = user['grado']
    session['seccion'] = user['seccion']

    registrar_log(user['cui'], 'login', 'Inicio de sesion exitoso', request.remote_addr)

    return jsonify({
        'success': True,
        'user': {
            'cui': user['cui'],
            'nombre_completo': user['nombre_completo'],
            'rol': user['rol'],
            'grado': user['grado'],
            'seccion': user['seccion']
        }
    })

@app.route('/api/logout', methods=['POST'])
@login_required
@csrf_required
def api_logout():
    cui = session.get('cui', 'desconocido')
    registrar_log(cui, 'logout', 'Cierre de sesion del usuario', request.remote_addr)
    session.clear()
    return jsonify({'success': True, 'mensaje': 'Sesión cerrada correctamente.'})

@app.route('/api/session', methods=['GET'])
@login_required
def api_session():
    return jsonify({
        'cui': session['cui'],
        'nombre_completo': session['nombre'],
        'rol': session['rol'],
        'grado': session['grado'],
        'seccion': session['seccion'],
        'csrf_token': generar_csrf_token()
    })

@app.route('/api/papeleria', methods=['GET'])
@login_required
def api_papeleria():
    cui = session.get('cui')
    rol = session.get('rol')

    if rol == 'admin':
        target_cui = request.args.get('cui', cui)
    else:
        target_cui = cui

    tipos = obtener_tipos()

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT cui, nombre_completo, grado, seccion, anios FROM usuarios WHERE cui = ?", (target_cui,))
    row = cursor.fetchone()
    conn.close()

    estudiante = dict(row) if row else None
    anios_estudiante = []
    if estudiante and estudiante.get('anios'):
        anios_estudiante = [int(a.strip()) for a in estudiante['anios'].split(',') if a.strip().isdigit()]

    if not anios_estudiante:
        anios_estudiante = ANIOS

    anio_filtro = request.args.get('anio')
    if anio_filtro:
        try:
            anio_filtro = int(anio_filtro)
        except (ValueError, TypeError):
            anio_filtro = None
        if anio_filtro and anio_filtro in anios_estudiante:
            anios_estudiante = [anio_filtro]

    registros = obtener_papeleria_estudiante(target_cui)

    registro_map = {}
    for r in registros:
        registro_map[f"{r['anio']}_{r['tipo']}"] = r

    grid = []
    for anio in anios_estudiante:
        fila = {'anio': anio, 'items': []}
        for t in tipos:
            key = f"{anio}_{t['nombre']}"
            if key in registro_map:
                r = registro_map[key]
                fila['items'].append({
                    'tipo': t['nombre'],
                    'estado': r['estado'],
                    'observaciones': r.get('observaciones', ''),
                    'id': r['id']
                })
            else:
                fila['items'].append({
                    'tipo': t['nombre'],
                    'estado': None,
                    'observaciones': '',
                    'id': None
                })
        grid.append(fila)

    registrar_log(cui, 'consulta_papeleria', f"Consulto papelería de {target_cui}", request.remote_addr)

    return jsonify({
        'tipos': [t['nombre'] for t in tipos],
        'anios': anios_estudiante,
        'anio_actual': anio_filtro or (anios_estudiante[0] if anios_estudiante else None),
        'grid': grid,
        'estudiante': estudiante
    })

@app.route('/api/admin/papeleria', methods=['POST'])
@csrf_required
@admin_required
def api_admin_actualizar_papeleria():
    data = request.get_json() or {}
    cui = data.get('cui')
    anio = data.get('anio')
    tipo = data.get('tipo')
    estado = data.get('estado')
    observaciones = data.get('observaciones', '')

    if not all([cui, anio, tipo, estado]):
        return jsonify({'error': 'Faltan campos obligatorios.'}), 400

    if estado not in ESTADOS:
        return jsonify({'error': 'Estado inválido.'}), 400

    exito = actualizar_papeleria(cui, anio, tipo, estado, observaciones, updated_by=session.get('cui'))
    if exito:
        registrar_log(
            session.get('cui'),
            'actualizar_papeleria',
            f"Actualizó papelería de {cui}: {anio} - {tipo} = {estado}",
            request.remote_addr
        )
        return jsonify({'success': True, 'mensaje': 'Estado actualizado correctamente.'})
    else:
        return jsonify({'error': 'Error al actualizar.'}), 500

@app.route('/api/admin/buscar', methods=['GET'])
@admin_required
def api_admin_buscar():
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify([])
    resultados = buscar_estudiantes(q)
    return jsonify(resultados)

@app.route('/api/admin/alumnos', methods=['GET', 'POST'])
@csrf_required
@admin_required
def api_admin_alumnos():
    if request.method == 'GET':
        alumnos = obtener_todos_usuarios()
        for al in alumnos:
            exp = obtener_expediente(al['cui'])
            if exp:
                al['expediente'] = exp
        return jsonify(alumnos)

    data = request.get_json() or {}
    cui = data.get('cui')
    nombre_completo = data.get('nombre_completo')
    grado = data.get('grado')
    seccion = data.get('seccion')
    anios = data.get('anios', '')
    password = data.get('password')
    caja = data.get('caja', '')
    ubicacion = data.get('ubicacion', '')
    posicion_caja = data.get('posicion_caja', '')

    if not all([cui, nombre_completo, grado, seccion]):
        return jsonify({'error': 'CUI, nombre, grado y sección son obligatorios.'}), 400

    if anios:
        anios_list = [a.strip() for a in anios.split(',') if a.strip()]
        for a in anios_list:
            if not a.isdigit():
                return jsonify({'error': f'Año inválido: {a}'}), 400
        if len(anios_list) > 3:
            return jsonify({'error': 'Máximo 3 años permitidos.'}), 400

    if not password:
        password = cui

    exito = crear_usuario(
        cui=cui,
        nombre=nombre_completo,
        rol='estudiante',
        grado=grado,
        seccion=seccion,
        anios=anios,
        password=password,
        acepto_consentimiento=False
    )

    if exito:
        registrar_log(
            session.get('cui'),
            'crear_estudiante',
            f"Registró al estudiante {nombre_completo} con CUI {cui}",
            request.remote_addr
        )
        if caja or ubicacion:
            crear_expediente(cui, caja, ubicacion, posicion_caja, session.get('cui'))
        exp = obtener_expediente(cui)
        return jsonify({'success': True, 'mensaje': 'Estudiante registrado correctamente.', 'expediente': exp}), 201
    else:
        return jsonify({'error': 'El CUI ya se encuentra registrado en el sistema.'}), 400

@app.route('/api/admin/alumnos/<cui>', methods=['DELETE'])
@csrf_required
@admin_required
def api_admin_eliminar_alumno(cui):
    if cui == session.get('cui'):
        return jsonify({'error': 'No puede eliminarse a sí mismo.'}), 400

    exito = eliminar_usuario(cui)
    if exito:
        registrar_log(
            session.get('cui'),
            'eliminar_estudiante',
            f"Eliminó al estudiante con CUI {cui} y su papelería asociada",
            request.remote_addr
        )
        return jsonify({'success': True, 'mensaje': 'Estudiante eliminado correctamente.'})
    else:
        return jsonify({'error': 'Estudiante no encontrado.'}), 404

@app.route('/api/admin/alumnos/<cui>', methods=['PUT'])
@csrf_required
@admin_required
def api_admin_actualizar_alumno(cui):
    data = request.get_json() or {}
    nombre_completo = data.get('nombre_completo')
    grado = data.get('grado')
    seccion = data.get('seccion')
    anios = data.get('anios')
    caja = data.get('caja')
    ubicacion = data.get('ubicacion')
    posicion_caja = data.get('posicion_caja')

    if anios:
        anios_list = [a.strip() for a in anios.split(',') if a.strip()]
        for a in anios_list:
            if not a.isdigit():
                return jsonify({'error': f'Año inválido: {a}'}), 400
        if len(anios_list) > 3:
            return jsonify({'error': 'Máximo 3 años permitidos.'}), 400

    exito = actualizar_usuario(cui, nombre_completo, grado, seccion, anios)
    if exito:
        registrar_log(
            session.get('cui'),
            'actualizar_estudiante',
            f"Actualizó datos del estudiante CUI {cui}",
            request.remote_addr
        )
        if caja is not None or ubicacion is not None or posicion_caja is not None:
            exp = obtener_expediente(cui)
            if exp:
                actualizar_expediente(cui, caja=caja, ubicacion=ubicacion,
                                       posicion_caja=posicion_caja, usuario_cui=session.get('cui'))
            else:
                crear_expediente(cui, caja or '', ubicacion or '', posicion_caja or '', session.get('cui'))
        return jsonify({'success': True, 'mensaje': 'Estudiante actualizado correctamente.'})
    else:
        return jsonify({'error': 'Estudiante no encontrado o sin cambios.'}), 404

@app.route('/api/admin/alumnos/exportar', methods=['GET'])
@admin_required
def api_admin_exportar_alumnos():
    q = request.args.get('q', '').strip()

    conn = get_connection()
    cursor = conn.cursor()
    if q:
        escaped = q.replace('%', '\\%').replace('_', '\\_')
        like = f"%{escaped}%"
        cursor.execute("""
            SELECT cui, nombre_completo, grado, seccion, anios, acepto_consentimiento
            FROM usuarios
            WHERE rol = 'estudiante' AND (nombre_completo LIKE ? ESCAPE '\\' OR cui LIKE ? ESCAPE '\\')
            ORDER BY nombre_completo ASC
        """, (like, like))
    else:
        cursor.execute("""
            SELECT cui, nombre_completo, grado, seccion, anios, acepto_consentimiento
            FROM usuarios
            WHERE rol = 'estudiante'
            ORDER BY nombre_completo ASC
        """)
    rows = cursor.fetchall()
    conn.close()

    if not OPENPYXL_DISPONIBLE:
        return jsonify({'error': 'openpyxl no está instalado.'}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estudiantes"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['No.', 'CUI', 'Nombre Completo', 'Grado', 'Sección', 'Año(s)', 'Consentimiento']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, row in enumerate(rows, 1):
        consent = "Sí" if row['acepto_consentimiento'] else "No"
        values = [i, row['cui'], row['nombre_completo'], row['grado'], row['seccion'], row['anios'] or '', consent]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=i + 1, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 14

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    now = datetime.now()
    fecha_hora = now.strftime('%Y-%m-%d_%H-%M-%S')
    filename = f"alumnos_{fecha_hora}.xlsx"

    registrar_log(session.get('cui'), 'exportar_excel_alumnos', f"Exportó {len(rows)} estudiantes a Excel", request.remote_addr)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

@app.route('/api/admin/alumnos/importar', methods=['POST'])
@csrf_required
@admin_required
def api_admin_importar_alumnos():
    if 'archivo' not in request.files:
        return jsonify({'error': 'No se envió ningún archivo.'}), 400

    archivo = request.files['archivo']
    if not archivo.filename:
        return jsonify({'error': 'Archivo vacío.'}), 400

    accion = request.form.get('accion', 'preview')

    ext = archivo.filename.rsplit('.', 1)[-1].lower() if '.' in archivo.filename else ''
    if ext not in ('xlsx', 'csv', 'txt', 'docx'):
        return jsonify({'error': 'Formato no soportado. Use .xlsx, .csv, .txt o .docx'}), 400

    try:
        bytes_archivo = archivo.read()
        registros_brutos = parsear_archivo(bytes_archivo, archivo.filename, es_bytes=True)
        registros = normalizar_registros(registros_brutos)
    except Exception as e:
        return jsonify({'error': f'Error al leer el archivo: {str(e)}'}), 400

    if not registros:
        return jsonify({'error': 'No se pudo extraer información del archivo. Verifique que tenga columnas CUI, Nombre, Grado, Sección.'}), 400

    if accion == 'preview':
        return jsonify({
            'total': len(registros),
            'registros': registros[:50],
            'columnas_detectadas': list(registros[0].keys()) if registros else []
        })

    if accion == 'importar':
        creados = 0
        errores = []
        for reg in registros:
            cui = reg.get('cui', '').strip()
            nombre = reg.get('nombre_completo', '').strip()
            grado = reg.get('grado', '').strip()
            seccion = reg.get('seccion', '').strip()
            anios = reg.get('anios', '').strip()

            if not re.match(r'^\d{13}$', cui):
                errores.append(f'CUI inválido para "{nombre or cui}": debe tener 13 dígitos')
                continue
            if not nombre:
                errores.append(f'Falta nombre para CUI {cui}')
                continue

            existing = obtener_usuario_por_cui(cui)
            if existing:
                actualizar_usuario(cui, nombre, grado, seccion, anios or existing.get('anios', ''))
                creados += 1
            else:
                ok = crear_usuario(cui, nombre, 'estudiante', grado, seccion, cui, False, anios)
                if ok:
                    creados += 1
                else:
                    errores.append(f'Error al crear CUI {cui}')

        mensaje = f'✅ {creados} estudiante(s) procesados.'
        if errores:
            mensaje += f' ⚠️ {len(errores)} error(es): {" / ".join(errores[:5])}'
            if len(errores) > 5:
                mensaje += f' (+{len(errores)-5} más)'

        registrar_log(session.get('cui'), 'importar_alumnos', f"Importó {creados} estudiantes desde archivo ({len(errores)} errores)", request.remote_addr)
        return jsonify({'success': True, 'creados': creados, 'errores': errores, 'mensaje': mensaje})

    return jsonify({'error': 'Acción no válida.'}), 400

@app.route('/api/admin/tipos', methods=['GET', 'POST'])
@csrf_required
@admin_required
def api_admin_tipos():
    if request.method == 'GET':
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM tipos_papeleria WHERE activo = 1 ORDER BY categoria ASC, id ASC")
        rows = cursor.fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])

    data = request.get_json() or {}
    nombre = data.get('nombre', '').strip()
    categoria = data.get('categoria', 'General')
    if not nombre:
        return jsonify({'error': 'El nombre del tipo es obligatorio.'}), 400

    tipo_id = crear_tipo(nombre, categoria)
    if tipo_id:
        registrar_log(session.get('cui'), 'crear_tipo', f"Creó tipo de papelería: {nombre} [{categoria}]", request.remote_addr)
        return jsonify({'success': True, 'id': tipo_id, 'nombre': nombre, 'categoria': categoria}), 201
    else:
        return jsonify({'error': 'El tipo ya existe.'}), 400

@app.route('/api/admin/tipos/count/<int:tipo_id>', methods=['GET'])
@admin_required
def api_admin_tipos_count(tipo_id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT nombre FROM tipos_papeleria WHERE id = ?", (tipo_id,))
    tipo = cursor.fetchone()
    if not tipo:
        conn.close()
        return jsonify({'error': 'Tipo no encontrado.'}), 404
    cursor.execute("SELECT COUNT(*) as count FROM papeleria WHERE tipo = ?", (tipo['nombre'],))
    count = cursor.fetchone()['count']
    conn.close()
    return jsonify({'affected': count})

@app.route('/api/admin/tipos/<int:tipo_id>', methods=['PUT', 'DELETE'])
@csrf_required
@admin_required
def api_admin_tipo_detail(tipo_id):
    if request.method == 'PUT':
        data = request.get_json() or {}
        nombre = data.get('nombre', '').strip()
        categoria = data.get('categoria')
        if not nombre:
            return jsonify({'error': 'El nombre es obligatorio.'}), 400
        exito = actualizar_tipo(tipo_id, nombre, categoria)
        if exito:
            registrar_log(session.get('cui'), 'editar_tipo', f"Editó tipo de papelería ID {tipo_id} a: {nombre}", request.remote_addr)
            return jsonify({'success': True})
        else:
            return jsonify({'error': 'Tipo no encontrado.'}), 404

    result = eliminar_tipo(tipo_id)
    if isinstance(result, dict) and result.get('deleted'):
        detalles = f"Eliminó tipo de papelería ID {tipo_id}"
        if result['affected_records'] > 0:
            detalles += f" (se eliminaron {result['affected_records']} registro(s) de papelería asociados)"
        registrar_log(session.get('cui'), 'eliminar_tipo', detalles, request.remote_addr)
        return jsonify({'success': True, 'affected_records': result['affected_records']})
    else:
        return jsonify({'error': 'Tipo no encontrado.'}), 404

@app.route('/api/admin/tipos/categorias', methods=['GET'])
@admin_required
def api_admin_categorias():
    return jsonify(CATEGORIAS)

@app.route('/api/admin/logs', methods=['GET'])
@admin_required
def api_admin_obtener_logs():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    per_page = min(per_page, 200)
    offset = (page - 1) * per_page

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) as total FROM logs")
    total = cursor.fetchone()['total']
    cursor.execute("""
        SELECT l.*, u.nombre_completo AS usuario_nombre
        FROM logs l
        LEFT JOIN usuarios u ON l.usuario_cui = u.cui
        ORDER BY l.id DESC
        LIMIT ? OFFSET ?
    """, (per_page, offset))
    rows = cursor.fetchall()
    conn.close()

    logs = [dict(r) for r in rows]
    return jsonify({'logs': logs, 'total': total, 'page': page, 'per_page': per_page})

@app.route('/api/admin/stats', methods=['GET'])
@admin_required
def api_admin_stats():
    conn = get_connection()
    cursor = conn.cursor()
    
    cursor.execute("SELECT COUNT(*) as count FROM usuarios WHERE rol = 'estudiante'")
    total_estudiantes = cursor.fetchone()['count']
    
    cursor.execute("SELECT COUNT(*) as count FROM tipos_papeleria WHERE activo = 1")
    total_tipos = cursor.fetchone()['count']
    
    cursor.execute("""
        SELECT u.cui, u.anios,
               COUNT(p.id) as total_docs,
               SUM(CASE WHEN p.estado = 'en_orden' THEN 1 ELSE 0 END) as en_orden_count
        FROM usuarios u
        LEFT JOIN papeleria p ON u.cui = p.estudiante_cui
        WHERE u.rol = 'estudiante'
        GROUP BY u.cui
    """)
    rows = cursor.fetchall()
    conn.close()
    
    completos = 0
    incompletos = 0
    
    for row in rows:
        anios_str = row['anios'] or ''
        anios = [a.strip() for a in anios_str.split(',') if a.strip()]
        if not anios:
            incompletos += 1
            continue
        req_count = len(anios) * total_tipos
        if req_count == 0:
            incompletos += 1
            continue
        if (row['en_orden_count'] or 0) >= req_count:
            completos += 1
        else:
            incompletos += 1
    
    return jsonify({
        'total_estudiantes': total_estudiantes,
        'total_tipos': total_tipos,
        'completos': completos,
        'incompletos': incompletos
    })

@app.route('/api/admin/papeleria/export', methods=['GET'])
@admin_required
def api_admin_export_csv():
    grado = request.args.get('grado', '')
    seccion = request.args.get('seccion', '')
    estado_filtro = request.args.get('estado', '')
    anio_filtro = request.args.get('anio', '')

    conn = get_connection()
    cursor = conn.cursor()
    sql = """
        SELECT u.cui, u.nombre_completo, u.grado, u.seccion,
               p.anio, p.tipo, p.estado, p.observaciones, p.updated_at
        FROM papeleria p
        JOIN usuarios u ON p.estudiante_cui = u.cui
        WHERE u.rol = 'estudiante'
    """
    params = []
    if grado:
        sql += " AND u.grado = ?"
        params.append(grado)
    if seccion:
        sql += " AND u.seccion = ?"
        params.append(seccion)
    if estado_filtro:
        sql += " AND p.estado = ?"
        params.append(estado_filtro)
    if anio_filtro:
        try:
            anio_int = int(anio_filtro)
            sql += " AND p.anio = ?"
            params.append(anio_int)
        except (ValueError, TypeError):
            pass
    sql += " ORDER BY u.nombre_completo, p.anio, p.tipo"

    cursor.execute(sql, params)
    rows = cursor.fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['CUI', 'Nombre', 'Grado', 'Seccion', 'Anio', 'Tipo', 'Estado', 'Observaciones', 'Ultima_Actualizacion'])
    for r in rows:
        writer.writerow([r['cui'], r['nombre_completo'], r['grado'], r['seccion'],
                        r['anio'], r['tipo'], r['estado'], r['observaciones'], r['updated_at'] or ''])

    response = app.response_class(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=papeleria_export.csv'}
    )
    registrar_log(session.get('cui'), 'exportar_csv', 'Exportó papelería a CSV', request.remote_addr)
    return response

@app.route('/api/admin/exportar/grado', methods=['GET'])
@admin_required
def api_admin_exportar_grado():
    grado = request.args.get('grado', '').strip()
    seccion = request.args.get('seccion', '').strip()
    caja = request.args.get('caja', '').strip()

    if not grado or not seccion:
        return jsonify({'error': 'Grado y sección son obligatorios.'}), 400

    if not OPENPYXL_DISPONIBLE:
        return jsonify({'error': 'openpyxl no está instalado. Ejecute: pip install openpyxl'}), 500

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT cui, nombre_completo, grado, seccion
        FROM usuarios
        WHERE rol = 'estudiante' AND grado = ? AND seccion = ?
        ORDER BY nombre_completo ASC
    """, (grado, seccion))
    rows = cursor.fetchall()
    conn.close()

    if not rows:
        return jsonify({'error': 'No se encontraron estudiantes para ese grado y sección.'}), 404

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{grado} - {seccion}"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ['No.', 'CUI', 'Nombre Completo', 'Grado', 'Sección']
    if caja:
        headers.append('Caja')

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, row in enumerate(rows, 1):
        values = [i, row['cui'], row['nombre_completo'], row['grado'], row['seccion']]
        if caja:
            values.append(caja)
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=i + 1, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    if caja:
        ws.column_dimensions['F'].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    now = datetime.now()
    fecha = now.strftime('%Y-%m-%d')
    hora = now.strftime('%H-%M-%S')
    nivel = "basicos" if any(x in grado.lower() for x in ['basico', 'básico']) else "primaria" if any(x in grado.lower() for x in ['primaria']) else "carrera"
    grado_clean = grado.replace(' ', '_').replace('/', '-')
    seccion_clean = seccion.replace(' ', '_').replace('/', '-')
    filename = f"{nivel}_{grado_clean}_{seccion_clean}_{fecha}_{hora}.xlsx"
    if caja:
        caja_clean = caja.replace(' ', '_').replace('/', '-')
        filename = f"{nivel}_{grado_clean}_{seccion_clean}_{caja_clean}_{fecha}_{hora}.xlsx"

    registrar_log(
        session.get('cui'), 'exportar_excel',
        f"Exportó {len(rows)} estudiantes de {grado} {seccion}" + (f" (Caja: {caja})" if caja else ""),
        request.remote_addr
    )
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

@app.route('/api/admin/papeleria/alertas', methods=['GET'])
@admin_required
def api_admin_alertas():
    estudiantes = obtener_estudiantes_incompletos(limit=30)
    return jsonify(estudiantes)

@app.route('/api/admin/papeleria/buscar', methods=['GET'])
@admin_required
def api_admin_buscar_papeleria():
    estado = request.args.get('estado', '').strip()
    grado = request.args.get('grado', '').strip()
    seccion = request.args.get('seccion', '').strip()
    if not estado or estado not in ESTADOS:
        return jsonify({'error': 'Estado inválido'}), 400
    resultados = buscar_papeleria_por_estado(estado, grado or None, seccion or None)
    return jsonify(resultados)

@app.route('/api/admin/papeleria/batch', methods=['POST'])
@csrf_required
@admin_required
def api_admin_papeleria_batch():
    data = request.get_json() or {}
    cui = data.get('cui')
    anios = data.get('anios', [])
    tipo = data.get('tipo')
    estado = data.get('estado')
    observaciones = data.get('observaciones', '')

    if not all([cui, anios, tipo, estado]):
        return jsonify({'error': 'Faltan campos obligatorios.'}), 400
    if estado not in ESTADOS:
        return jsonify({'error': 'Estado inválido.'}), 400

    ok = actualizar_papeleria_batch(cui, anios, tipo, estado, observaciones, session.get('cui'))
    registrar_log(
        session.get('cui'), 'batch_papeleria',
        f"Actualización batch de {cui}: {len(anios)} año(s) - {tipo} = {estado} ({ok} ok)",
        request.remote_addr
    )
    return jsonify({'success': True, 'actualizados': ok})

@app.route('/api/admin/stats/grados', methods=['GET'])
@admin_required
def api_admin_stats_grados():
    stats = obtener_stats_por_grado()
    return jsonify(stats)


# ── Expediente Físico ──

@app.route('/api/admin/expediente/<cui>', methods=['GET'])
@admin_required
def api_admin_obtener_expediente(cui):
    exp = obtener_expediente(cui)
    if not exp:
        return jsonify({'error': 'Sin expediente físico registrado.'}), 404
    return jsonify(exp)


@app.route('/api/admin/expediente/<cui>', methods=['PUT'])
@csrf_required
@admin_required
def api_admin_guardar_expediente(cui):
    data = request.get_json() or {}
    exp = crear_expediente(cui, data.get('caja', ''), data.get('ubicacion', ''),
                           data.get('posicion_caja', ''), session.get('cui'))
    if not exp:
        exp = actualizar_expediente(cui, caja=data.get('caja'), ubicacion=data.get('ubicacion'),
                                     posicion_caja=data.get('posicion_caja'),
                                     usuario_cui=session.get('cui'))
    registrar_log(session.get('cui'), 'guardar_expediente', f"Guardó expediente físico para CUI {cui}", request.remote_addr)
    return jsonify(exp or {'error': 'Error al guardar expediente.'}), 200 if exp else 500


@app.route('/api/admin/expediente/<cui>/verificar', methods=['POST'])
@csrf_required
@admin_required
def api_admin_verificar_expediente(cui):
    exp = registrar_verificacion(cui, session.get('cui'))
    if not exp:
        return jsonify({'error': 'El estudiante no tiene expediente físico registrado.'}), 404
    registrar_log(session.get('cui'), 'verificar_expediente', f"Verificó físicamente expediente de CUI {cui}", request.remote_addr)
    return jsonify({'success': True, 'expediente': exp})


@app.route('/api/admin/expediente/<cui>/estado', methods=['POST'])
@csrf_required
@admin_required
def api_admin_cambiar_estado_expediente(cui):
    data = request.get_json() or {}
    estado = data.get('estado')
    notas = data.get('notas', '')
    if estado not in ESTADOS_FISICOS:
        return jsonify({'error': f'Estado inválido. Use: {", ".join(ESTADOS_FISICOS)}'}), 400
    exp = cambiar_estado_expediente(cui, estado, session.get('cui'), notas)
    if not exp:
        return jsonify({'error': 'El estudiante no tiene expediente físico registrado.'}), 404
    registrar_log(session.get('cui'), f'estado_expediente_{estado.lower()}', f"Expediente de CUI {cui} marcado como {estado}", request.remote_addr)
    return jsonify({'success': True, 'expediente': exp})


@app.route('/api/admin/expediente/<cui>/movimientos', methods=['GET'])
@admin_required
def api_admin_movimientos_expediente(cui):
    movs = obtener_movimientos(cui)
    return jsonify(movs)


@app.route('/api/admin/expedientes/extraviados', methods=['GET'])
@admin_required
def api_admin_expedientes_extraviados():
    return jsonify(obtener_extraviados())


@app.route('/api/admin/expedientes/stats-grados', methods=['GET'])
@admin_required
def api_admin_stats_expedientes():
    return jsonify(obtener_expedientes_por_grado())

@app.route('/api/admin/debug/stats-grados', methods=['GET'])
@admin_required
def api_admin_debug_stats_grados():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) as cnt FROM tipos_papeleria WHERE activo = 1")
    total_tipos = cursor.fetchone()['cnt']
    cursor.execute("""
        SELECT u.grado, u.seccion, u.cui, u.nombre_completo, u.anios,
               SUM(CASE WHEN p.estado = 'en_orden' THEN 1 ELSE 0 END) as en_orden,
               SUM(CASE WHEN p.estado = 'no_entregado' THEN 1 ELSE 0 END) as no_entregado,
               SUM(CASE WHEN p.estado = 'hace_falta' THEN 1 ELSE 0 END) as hace_falta,
               COUNT(p.id) as registros,
               CASE WHEN u.anios IS NULL OR u.anios = '' THEN 0
                    ELSE (LENGTH(u.anios) - LENGTH(REPLACE(u.anios, ',', '')) + 1)
               END * ? as requeridos
        FROM usuarios u
        LEFT JOIN papeleria p ON u.cui = p.estudiante_cui
        WHERE u.rol = 'estudiante'
        GROUP BY u.cui
        ORDER BY u.grado, u.seccion, u.nombre_completo
    """, (total_tipos,))
    rows = cursor.fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/admin/grados-secciones', methods=['GET'])
@admin_required
def api_admin_grados_secciones():
    grados, secciones = obtener_grados_secciones()
    return jsonify({'grados': grados, 'secciones': secciones})

# ── QR / Verificación ──

@app.route('/verificar/<token>')
def verificar_documento(token):
    info = verificar_por_token(token)
    if not info:
        return render_template('verificar.html', valido=False, error='Token de verificación inválido o expirado.')
    return render_template('verificar.html', valido=True, estudiante=info)

@app.route('/api/verificacion', methods=['GET'])
@login_required
def api_verificacion_token():
    cui = session.get('cui')
    token = obtener_token_verificacion(cui)
    if not token:
        return jsonify({'error': 'No se pudo generar el token de verificación.'}), 500
    url = f"{request.host_url}verificar/{token}"
    qr_data = None
    if QR_DISPONIBLE:
        try:
            img = qrcode.make(url)
            buf = BytesIO()
            img.save(buf, format='PNG')
            b64 = base64.b64encode(buf.getvalue()).decode()
            qr_data = f"data:image/png;base64,{b64}"
        except Exception:
            pass
    return jsonify({'token': token, 'url': url, 'qr_data': qr_data})

@app.route('/api/admin/verificacion/<cui>', methods=['GET'])
@login_required
def api_admin_verificacion(cui):
    if session.get('rol') != 'admin' and session.get('cui') != cui:
        return jsonify({'error': 'No autorizado'}), 403
    token = obtener_token_verificacion(cui)
    if not token:
        return jsonify({'error': 'Estudiante no encontrado.'}), 404
    url = f"{request.host_url}verificar/{token}"
    qr_data = None
    if QR_DISPONIBLE:
        try:
            img = qrcode.make(url)
            buf = BytesIO()
            img.save(buf, format='PNG')
            b64 = base64.b64encode(buf.getvalue()).decode()
            qr_data = f"data:image/png;base64,{b64}"
        except Exception:
            pass
    return jsonify({'token': token, 'url': url, 'qr_data': qr_data})

@app.route('/api/cambiar-password', methods=['POST'])
@login_required
@csrf_required
def api_cambiar_password():
    data = request.get_json() or {}
    current = data.get('password_actual')
    new = data.get('password_nuevo')
    if not current or not new:
        return jsonify({'error': 'Ambos campos son obligatorios.'}), 400
    if len(new) < 4:
        return jsonify({'error': 'La nueva contraseña debe tener al menos 4 caracteres.'}), 400
    user = validar_usuario(session['cui'], current)
    if not user:
        return jsonify({'error': 'Contraseña actual incorrecta.'}), 401
    p_hash = hash_password(new)
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("UPDATE usuarios SET password_hash = ? WHERE cui = ?", (p_hash, session['cui']))
    conn.commit()
    conn.close()
    registrar_log(session['cui'], 'cambiar_password', 'Cambió su contraseña', request.remote_addr)
    return jsonify({'success': True, 'mensaje': 'Contraseña actualizada correctamente.'})

if __name__ == '__main__':
    debug_mode = os.environ.get('FLASK_DEBUG', '0') == '1'
    port = int(os.environ.get('PORT', 5000))

    # ── Obtener IP local ──
    local_ips = []
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('10.255.255.255', 1))
        local_ips.append(s.getsockname()[0])
        s.close()
    except Exception:
        pass
    try:
        import subprocess
        r = subprocess.run(['hostname', '-I'], capture_output=True, text=True, timeout=2)
        local_ips.extend(r.stdout.strip().split())
    except Exception:
        pass

    print(f"\n{'='*55}")
    print(f"  🚀 Servidor iniciado — Control de Papelería")
    print(f"  {'='*55}")
    print(f"  Local:    http://127.0.0.1:{port}")
    for ip in local_ips:
        if ip.strip():
            print(f"  Red:      http://{ip}:{port}")
    print(f"  {'='*55}")

    # ── Tunnel público con cloudflared (si está instalado) ──
    cloudflared_path = None
    for candidate in ['cloudflared', '/data/data/com.termux/files/usr/bin/cloudflared']:
        try:
            r = subprocess.run(['which', candidate], capture_output=True, text=True, timeout=2)
            if r.returncode == 0 and r.stdout.strip():
                cloudflared_path = r.stdout.strip()
                break
        except Exception:
            try:
                if os.path.exists(candidate):
                    cloudflared_path = candidate
                    break
            except Exception:
                pass

    if cloudflared_path:
        tunnel_urls = []

        def leer_tunnel(proc):
            for line in iter(proc.stdout.readline, ''):
                if not line:
                    break
                if 'trycloudflare.com' in line:
                    idx = line.find('https://')
                    if idx >= 0:
                        url = line[idx:].split()[0].rstrip('.')
                        tunnel_urls.append(url)
                        print(f"  🌍 Público:  {url}")
                        print(f"  {'='*55}\n")
                        generar_qr_terminal(url)

        def generar_qr_terminal(url):
            try:
                import qrcode
                qr = qrcode.QRCode(box_size=1, border=1)
                qr.add_data(url)
                qr.make(fit=True)
                print("  📷 Escanea el QR para abrir desde cualquier dispositivo:\n")
                for row in qr.get_matrix():
                    print('  ' + ''.join('█' if c else ' ' for c in row))
                print()
            except Exception:
                pass

        try:
            proc = subprocess.Popen(
                [cloudflared_path, 'tunnel', '--url', f'http://127.0.0.1:{port}',
                 '--no-autoupdate', '--protocol', 'http2'],
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, bufsize=1
            )
            hilo = threading.Thread(target=leer_tunnel, args=(proc,), daemon=True)
            hilo.start()
        except Exception as e:
            print(f"  ⚠️  cloudflared no pudo iniciar: {e}")
    else:
        print(f"  💡 Instala cloudflared para acceso público:")
        print(f"     pkg install cloudflared  (o https://developers.cloudflare.com/cloudflare-one/connections/connect-networks/downloads/)")
        print(f"  {'='*55}\n")

    app.run(host='0.0.0.0', port=port, debug=debug_mode)

